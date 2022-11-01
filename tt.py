
import openpyxl
 
path = "C:\\Users\\Admin\\Desktop\\demo.xlsx"
 
wb_obj = openpyxl.load_workbook('Order_Template.xlsx')
# wb2=wb_obj['Translator']
wb2= wb_obj.sheetnames[0]
# Get workbook active sheet objec0
# from the active attribute
sheet_obj = wb_obj.active
wb2= sheet_obj.sheetnames[0]
print(wb2)

cell_obj = sheet_obj.cell(row = 1, column = 3)
 
# Print value of cell object
# using the value attribute
print(cell_obj.value)